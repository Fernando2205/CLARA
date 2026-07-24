from __future__ import annotations

import base64
import csv
import html
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

SIN_CONTAR = "Sin contar"


def _pending_rows(connection: sqlite3.Connection, session_id: str, bodega: str) -> list[dict]:
    articulos = connection.execute(
        """
        SELECT sku, articulo, bodega, unidad, stock_sistema
        FROM articulos
        WHERE bodega = ?
          AND id NOT IN (
              SELECT articulo_id FROM registros WHERE sesion_id = ?
          )
        ORDER BY articulo
        """,
        (bodega, session_id),
    ).fetchall()
    return [
        {
            "sku": row["sku"],
            "articulo": row["articulo"],
            "bodega": row["bodega"],
            "unidad": row["unidad"],
            "stock_sistema": row["stock_sistema"],
            "cantidad_fisica": None,
            "estado_producto": SIN_CONTAR,
            "corregido": 0,
        }
        for row in articulos
    ]


def report_data(
    connection: sqlite3.Connection, session_id: str, alcance: str = "contados"
) -> tuple[sqlite3.Row, list]:
    session = connection.execute(
        """
        SELECT s.*, u.nombre, u.cargo, u.firma_path FROM sesiones s
        JOIN usuarios u ON u.id = s.usuario_id
        WHERE s.id = ?
        """,
        (session_id,),
    ).fetchone()
    if not session:
        raise ValueError("Sesión no encontrada")

    contados = connection.execute(
        """
        SELECT r.*, a.sku, a.articulo, a.stock_sistema, a.bodega
        FROM registros r JOIN articulos a ON a.id = r.articulo_id
        WHERE r.sesion_id = ? ORDER BY a.articulo
        """,
        (session_id,),
    ).fetchall()

    if alcance == "contados":
        return session, list(contados)

    pendientes = _pending_rows(connection, session_id, session["bodega"])
    if alcance == "faltantes":
        return session, pendientes
    return session, [*contados, *pendientes]


def _fisico(row) -> float | None:
    value = row["cantidad_fisica"]
    return None if value is None else value


def generate_csv(path: Path, records: list) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.writer(output, delimiter=";")
        writer.writerow(["CANTIDAD", "Nr.Artículo", "Artículo", "Unidad", "SD"])
        for row in records:
            fisico = _fisico(row)
            writer.writerow([
                fisico if fisico is not None else SIN_CONTAR, row["sku"] or "", row["articulo"],
                row["unidad"], fisico if fisico is not None else "",
            ])


def generate_xlsx(path: Path, session: sqlite3.Row, records: list) -> None:
    workbook = Workbook()
    detail = workbook.active
    detail.title = "Detalle"
    headers = [
        "Nr.Artículo", "Artículo", "Bodega", "Cantidad física", "Unidad",
        "Stock sistema", "Diferencia", "Estado", "Corregido",
    ]
    detail.append(headers)
    for row in records:
        fisico = _fisico(row)
        delta = fisico - row["stock_sistema"] if fisico is not None else None
        detail.append([
            row["sku"] or "", row["articulo"], row["bodega"],
            fisico if fisico is not None else SIN_CONTAR, row["unidad"], row["stock_sistema"],
            delta if delta is not None else "", row["estado_producto"] or "",
            "Sí" if row["corregido"] else "No",
        ])

    differences = workbook.create_sheet("Diferencias")
    differences.append(["Artículo", "Físico", "Sistema", "Delta"])
    for row in records:
        fisico = _fisico(row)
        delta = fisico - row["stock_sistema"] if fisico is not None else None
        differences.append([
            row["articulo"], fisico if fisico is not None else SIN_CONTAR,
            row["stock_sistema"], delta if delta is not None else "",
        ])

    for sheet in (detail, differences):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0069AA")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(45, max(len(str(cell.value or "")) for cell in column) + 2)
            sheet.column_dimensions[column[0].column_letter].width = width

    detail["K1"] = "Sesión"
    detail["K2"] = session["id"]
    detail["K3"] = session["nombre"]
    detail["K4"] = session["hash_firma"] or "Sin firma"
    workbook.save(path)


def generate_pdf(path: Path, session: sqlite3.Row, records: list) -> None:
    rows = []
    for row in records:
        fisico = _fisico(row)
        if fisico is None:
            rows.append(
                "<tr class='pendiente'>"
                f"<td>{html.escape(row['sku'] or '—')}</td>"
                f"<td>{html.escape(row['articulo'])}</td>"
                f"<td colspan='3'>{SIN_CONTAR}</td>"
                "</tr>"
            )
            continue
        delta = fisico - row["stock_sistema"]
        delta_class = "positive" if delta > 0 else "negative" if delta < 0 else ""
        rows.append(
            "<tr>"
            f"<td>{html.escape(row['sku'] or '—')}</td>"
            f"<td>{html.escape(row['articulo'])}</td>"
            f"<td>{fisico:g}</td>"
            f"<td>{html.escape(row['unidad'])}</td>"
            f"<td>{row['stock_sistema']:g}</td>"
            f"<td class='{delta_class}'>{delta:+g}</td>"
            "</tr>"
        )

    # Solo se estampa la firma real si la sesión quedó formalmente firmada
    # (hash_firma presente); un acta sin firmar no debe mostrar la firma personal.
    firma_path = (
        Path(session["firma_path"])
        if session["firma_path"] and session["hash_firma"]
        else None
    )
    firma_img_html = ""
    if firma_path and firma_path.exists():
        firma_data = base64.b64encode(firma_path.read_bytes()).decode()
        firma_img_html = f"<img class='firma-img' src='data:image/png;base64,{firma_data}' alt='Firma'>"

    document = f"""
    <!doctype html>
    <html lang="es">
    <meta charset="utf-8">
    <style>
      @page {{ size: A4 landscape; margin: 18mm; }}
      body {{ font: 12px Arial, sans-serif; color: #15324b; }}
      header {{ display:flex; justify-content:space-between; border-bottom:4px solid #ffcb05;
                padding-bottom:12px; margin-bottom:22px; }}
      h1 {{ color:#0069aa; margin:0 0 4px; font-size:24px; }}
      .brand {{ color:#0069aa; font-size:26px; font-weight:800; }}
      .meta {{ background:#eef7fc; padding:12px 16px; border-radius:10px; margin-bottom:18px; }}
      table {{ width:100%; border-collapse:collapse; }}
      th {{ background:#0069aa; color:white; text-align:left; padding:8px; }}
      td {{ padding:7px 8px; border-bottom:1px solid #d6e1e8; }}
      tr.pendiente td {{ color:#9a6500; font-style:italic; }}
      .positive {{ color:#218739; }} .negative {{ color:#c72d37; }}
      .signature {{ margin-top:24px; display:flex; align-items:flex-end; gap:24px;
                    border-top:1px solid #9eb4c3; padding-top:12px; }}
      .firma-img {{ height:60px; }}
      .hash {{ font-family:monospace; font-size:9px; word-break:break-all; color:#5d7180; }}
    </style>
    <body>
      <header><div><h1>Acta de toma física de inventario</h1>
      <span>Cuentas claras, cocina tranquila.</span></div><div class="brand">◆ CLARA</div></header>
      <section class="meta">
        <strong>{html.escape(session['bodega'])}</strong><br>
        Responsable: {html.escape(session['nombre'])} · {html.escape(session['cargo'])}<br>
        Inicio: {html.escape(session['inicio'])} · Cierre: {html.escape(session['fin'] or 'En curso')}
      </section>
      <table><thead><tr><th>SKU</th><th>Artículo</th><th>Físico</th><th>Unidad</th>
      <th>Sistema</th><th>Delta</th></tr></thead><tbody>{''.join(rows)}</tbody></table>
      <section class="signature">
        <div>
          <strong>Firma digital de la sesión</strong>
          <p class="hash">{html.escape(session['hash_firma'] or 'Sesión aún no firmada')}</p>
        </div>
        {firma_img_html}
      </section>
    </body></html>
    """
    try:
        from weasyprint import HTML
        HTML(string=document).write_pdf(path)
    except (ImportError, OSError):
        # macOS suele no incluir Pango/GLib. ReportLab mantiene el endpoint
        # operativo; en producción WeasyPrint conserva la plantilla completa.
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        styles = getSampleStyleSheet()
        story = [
            Paragraph("CLARA · Acta de toma física de inventario", styles["Title"]),
            Paragraph(html.escape(session["bodega"]), styles["Heading2"]),
            Paragraph(
                f"Responsable: {html.escape(session['nombre'])} · "
                f"{html.escape(session['cargo'])}",
                styles["BodyText"],
            ),
            Spacer(1, 5 * mm),
        ]
        data = [["SKU", "Artículo", "Físico", "Unidad", "Sistema", "Delta"]]
        for row in records:
            fisico = _fisico(row)
            if fisico is None:
                data.append([row["sku"] or "—", row["articulo"], SIN_CONTAR, "", "", ""])
                continue
            delta = fisico - row["stock_sistema"]
            data.append([
                row["sku"] or "—", row["articulo"], f"{fisico:g}",
                row["unidad"], f"{row['stock_sistema']:g}", f"{delta:+g}",
            ])
        table = Table(data, repeatRows=1, colWidths=[24 * mm, 78 * mm, 20 * mm, 25 * mm, 22 * mm, 20 * mm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0069AA")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D6E1E8")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F8FB")]),
        ]))
        story.extend([
            table,
            Spacer(1, 7 * mm),
            Paragraph(
                f"Firma digital: {html.escape(session['hash_firma'] or 'Sesión aún no firmada')}",
                styles["Code"],
            ),
        ])
        if firma_path and firma_path.exists():
            reader = ImageReader(str(firma_path))
            natural_w, natural_h = reader.getSize()
            max_w, max_h = 50 * mm, 20 * mm
            scale = min(max_w / natural_w, max_h / natural_h)
            story.extend([
                Spacer(1, 3 * mm),
                Image(str(firma_path), width=natural_w * scale, height=natural_h * scale),
            ])
        SimpleDocTemplate(
            str(path), pagesize=landscape(A4), rightMargin=12 * mm,
            leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        ).build(story)


def generate_reports(
    connection: sqlite3.Connection,
    session_id: str,
    formats: list[str],
    output_root: Path,
    alcance: str = "contados",
) -> dict[str, Path]:
    session, records = report_data(connection, session_id, alcance)
    output_dir = output_root / session_id
    output_dir.mkdir(parents=True, exist_ok=True)
    generated: dict[str, Path] = {}
    for file_format in dict.fromkeys(formats):
        path = output_dir / f"CLARA_{session_id[:8]}.{file_format}"
        if file_format == "pdf":
            generate_pdf(path, session, records)
        elif file_format == "xlsx":
            generate_xlsx(path, session, records)
        elif file_format == "csv":
            generate_csv(path, records)
        generated[file_format] = path
    return generated
